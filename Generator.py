import openpyxl as xl

# SETTINGS!!!

workbook_name = "PROF EMAIL LIST.xlsx" # Excel file with all the info, leave alone if you did not rename
intro_line_replace = "[Name of Professor]" # Thing to be replaced with prof. name
font = "Times New Roman" # Font of output stuffs, change to any supported excel font


# Initiates stuffs
wb = xl.load_workbook(workbook_name)
inputs = wb["INPUT"]
outputs = wb["OUTPUT"]
email_draft = inputs.cell(4,5).value
prof_emails = []
prof_names = []

def info_into_python(): # Import info from INPUT from excel sheet, stops as soon as something is wrong
    row_in = 4
    while True:
        if inputs.cell(row_in,2).value == None or inputs.cell(row_in,3).value == None:
            print(f"Imported all professor information up until row {row_in} exclusive, make sure this is what is intended!")
            break
        else:
            prof_emails.append(inputs.cell(row_in,3).value)
            prof_names.append(inputs.cell(row_in, 2).value)
        row_in += 1

def email_generator(professor_name): # Generates the email by replacing a name
    duped_email = email_draft
    duped_email = duped_email.replace(intro_line_replace, professor_name)
    return duped_email

info_into_python()

row_out = 4

for professor in prof_names: # Actually generates a
    Last_name = professor.split(" ")[len(professor.split(" ")) - 1]
    generated_email = email_generator(Last_name)
    outputs.cell(row_out,3).value = generated_email
    outputs.cell(row_out, 3).font = font
    outputs.cell(row_out, 2).value = prof_emails[prof_names.index(professor)]
    outputs.cell(row_out, 2).font = font
    wb.save()





